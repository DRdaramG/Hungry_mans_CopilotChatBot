"""
Utilities for processing file attachments.

Supported types
---------------
* **Images** (.jpg / .jpeg / .png / .gif / .webp)
  Read as raw bytes and base64-encoded for inclusion in a multimodal message.
* **CSV** (.csv)
  Parsed with the standard :mod:`csv` module and returned as formatted text.
* **Excel** (.xls / .xlsx)
  Parsed with *openpyxl* (preferred) or *xlrd* as a fallback; returned as
  formatted text. If neither library is installed the file name is reported
  instead.
"""

import base64
import csv
import os
from pathlib import Path
from typing import TypedDict


class FileData(TypedDict):
    """Common structure returned by :func:`process_file`."""

    type: str        # "image" | "text" | "error"
    data: str        # base64 string for images, text content otherwise
    name: str        # original file name
    path: str        # absolute file path
    # image-only
    mime: str        # e.g. "image/png"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_IMAGE_MIME: dict[str, str] = {
    ".jpg":  "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png":  "image/png",
    ".gif":  "image/gif",
    ".webp": "image/webp",
}


def _read_image(file_path: str) -> tuple[str, str]:
    """Return *(base64_data, mime_type)* for an image file."""
    with open(file_path, "rb") as fh:
        raw = fh.read()
    ext = Path(file_path).suffix.lower()
    mime = _IMAGE_MIME.get(ext, "image/png")
    return base64.b64encode(raw).decode("utf-8"), mime


def _read_csv(file_path: str) -> str:
    """Return CSV content as a fenced code block string."""
    with open(file_path, "r", encoding="utf-8", errors="replace") as fh:
        rows = list(csv.reader(fh))
    if not rows:
        return f"(empty CSV: {os.path.basename(file_path)})"
    body = "\n".join(",".join(cell for cell in row) for row in rows)
    return (
        f"[CSV File: {os.path.basename(file_path)}]\n"
        f"```csv\n{body}\n```"
    )


def _read_excel(file_path: str) -> str:
    """Return Excel content as a fenced code block string."""
    basename = os.path.basename(file_path)
    try:
        import openpyxl  # type: ignore

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sections: list[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = [
                ",".join(
                    "" if cell is None else str(cell) for cell in row
                )
                for row in ws.iter_rows(values_only=True)
            ]
            sections.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
        wb.close()
        return (
            f"[Excel File: {basename}]\n"
            f"```csv\n" + "\n\n".join(sections) + "\n```"
        )
    except ImportError:
        pass

    try:
        import xlrd  # type: ignore

        wb = xlrd.open_workbook(file_path)
        sections = []
        for idx in range(wb.nsheets):
            ws = wb.sheet_by_index(idx)
            rows = [
                ",".join(str(ws.cell_value(r, c)) for c in range(ws.ncols))
                for r in range(ws.nrows)
            ]
            sections.append(f"[Sheet: {ws.name}]\n" + "\n".join(rows))
        return (
            f"[Excel File: {basename}]\n"
            f"```csv\n" + "\n\n".join(sections) + "\n```"
        )
    except ImportError:
        pass

    return (
        f"[Excel File: {basename}]\n"
        "(Install openpyxl or xlrd to read Excel files automatically.)"
    )


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def process_file(file_path: str) -> FileData:
    """
    Inspect *file_path* and return a :class:`FileData` dict.

    The ``"type"`` field is:
    * ``"image"``  — the ``"data"`` field holds base64-encoded bytes and
                     ``"mime"`` holds the MIME type.
    * ``"text"``   — the ``"data"`` field holds the extracted text content.
    * ``"error"``  — the ``"data"`` field holds an error description.
    """
    ext = Path(file_path).suffix.lower()
    basename = os.path.basename(file_path)
    base: FileData = {"name": basename, "path": file_path, "mime": "", "data": "", "type": ""}

    if ext in _IMAGE_MIME:
        b64, mime = _read_image(file_path)
        return {**base, "type": "image", "data": b64, "mime": mime}

    if ext == ".csv":
        return {**base, "type": "text", "data": _read_csv(file_path)}

    if ext in {".xls", ".xlsx"}:
        return {**base, "type": "text", "data": _read_excel(file_path)}

    # Fallback: try plain-text
    try:
        with open(file_path, "r", encoding="utf-8", errors="replace") as fh:
            text = fh.read()
        return {
            **base,
            "type": "text",
            "data": f"[File: {basename}]\n```\n{text}\n```",
        }
    except OSError as exc:
        return {**base, "type": "error", "data": f"Cannot read file: {exc}"}
